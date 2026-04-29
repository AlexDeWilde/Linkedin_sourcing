#!/usr/bin/env python3
"""
Stage 06 — LLM Scoring.
Reads .md files from 05-LLMfiltered/, scores each via Ollama using criteria
from 06-score_crit.txt, and runs fully unattended:

  - Streams scoring JSON live to console
  - Recalculates fit_score from the breakdown (LLM arithmetic ignored)
  - Appends a row to 06-listings_db.xlsx
  - Writes _SCORING.md alongside the triplet
  - Renames and moves .md + .url + _SCORING.md to 06-LLM_scored/

Listings whose description contains "no longer accepting applications" are
moved to 05-LLMfiltered/closed/ and skipped.

Naming convention:
  [ref_nr 4d]_[fit_score 3d]_[YYYYMMDD]_[city up to 7 chars]_[title]_[company]_[source]
"""

import json
import re
import shutil
import sys
import unicodedata
from datetime import date
from pathlib import Path

import requests
from openpyxl import load_workbook

# Force unbuffered output so streaming tokens appear live in the console
sys.stdout.reconfigure(write_through=True)

# ── Paths ───────────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
INPUT_DIR  = BASE_DIR / "05-LLMfiltered"
OUTPUT_DIR = BASE_DIR / "06-LLM_scored"
CRITERIA   = BASE_DIR / "06-score_crit.txt"
EXCEL_FILE = BASE_DIR / "06-listings_db.xlsx"

# ── Ollama config ───────────────────────────────────────────────────────────────
OLLAMA_URL  = "http://192.168.68.52:11434/api/chat"
NUM_CTX     = 49152
MAX_TOKENS  = 2048

def _load_model(stage: int) -> str:
    cfg = Path(__file__).parent / "_model_config.txt"
    default = None
    for raw in cfg.read_text(encoding="utf-8").splitlines():
        line = raw.split("#")[0].strip()
        if not line:
            continue
        model, _, rhs = line.partition("-")
        model, rhs = model.strip(), rhs.strip()
        if model == "default":
            default = rhs  # rhs is the fallback model name
            continue
        if rhs and any(s.strip() == str(stage) for s in rhs.split(",")):
            return model
    if default:
        return default
    raise RuntimeError(f"No model assigned to stage {stage} and no default in _model_config.txt")

MODEL = _load_model(6)

def _unload_all_models():
    base = OLLAMA_URL.replace("/api/chat", "")
    try:
        ps = requests.get(f"{base}/api/ps", timeout=10).json()
    except Exception:
        return  # Ollama unreachable — will surface properly at call time
    for m in ps.get("models", []):
        requests.post(f"{base}/api/generate",
                      json={"model": m["name"], "keep_alive": 0}, timeout=30)
        print(f"  Unloaded {m['name']} from Ollama.", flush=True)

TODAY = date.today()

# ── Excel column order (after city column was added in session 06 setup) ────────
# A:ref_nr  B:fit_score  C:date_published  D:city  E:tag  F:job_title
# G:company_name  H:source  I:date_found  J:status  K:report_produced
# L:processed  M:url


# ── Helpers ─────────────────────────────────────────────────────────────────────

def city_code(city: str) -> str:
    """Up to 7-char uppercase city code. REMOTE for remote roles."""
    c = city.strip()
    if not c or c.upper() in ("REMOTE", "RMTE", "FULLY REMOTE", "REMOTE WORK", "WORLDWIDE"):
        return "REMOTE"
    nfd = unicodedata.normalize("NFD", c)
    ascii_only = "".join(ch for ch in nfd if unicodedata.category(ch) != "Mn")
    compact = re.sub(r"\s+", "", ascii_only)
    return compact[:7].upper() if compact else "UNKNOWN"


_INVALID_CHARS = re.compile(r'[<>:"/\\|?*]')

def sanitize(text: str) -> str:
    """Remove filename-invalid chars, replace spaces with underscores."""
    text = _INVALID_CHARS.sub("", text)
    text = text.replace(" ", "_")
    return text.strip("_")


def extract_url(md_content: str) -> str:
    m = re.search(r'\*\*URL:\*\*\s*(https?://\S+)', md_content)
    return m.group(1).rstrip(")") if m else ""


def next_ref_nr() -> int:
    """Read Excel and return last ref_nr + 1. Handles both text '0277' and integer 277."""
    wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    ws = wb.active
    max_ref = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[0]
        if val is not None:
            try:
                max_ref = max(max_ref, int(str(val).strip()))
            except (ValueError, TypeError):
                pass
    wb.close()
    return max_ref + 1


def lookup_ref_nr_by_url(url: str) -> int | None:
    """Search Excel column M (url) for a matching URL. Return ref_nr int if found."""
    if not EXCEL_FILE.exists() or not url:
        return None
    wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 13 and str(row[12] or "").strip() == url.strip():
            ref_val = row[0]
            wb.close()
            if ref_val is not None:
                try:
                    return int(str(ref_val).strip())
                except (ValueError, TypeError):
                    pass
            return None
    wb.close()
    return None


def update_excel_row_for_rescore(ref_nr: int, result: dict) -> None:
    """Update the Excel row matching ref_nr with new score data."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ref_str   = str(ref_nr).zfill(4)
    score_str = f"{result.get('fit_score', 0):03d}" if result.get("fit_score") is not None else None
    date_pub  = str(result.get("date_published", TODAY.strftime("%Y%m%d")))

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value or "").strip().zfill(4) == ref_str:
            row[1].value = score_str
            row[1].number_format = "@"
            row[2].value = format_date_dmy(date_pub)
            row[3].value = result.get("city_code", "")
            if result.get("client_company"):
                row[4].value = result.get("client_company")
            row[5].value = result.get("job_title", "")
            row[6].value = result.get("company_name", "")
            row[7].value = result.get("source", "LinkedIn")
            row[8].value = TODAY.strftime("%d/%m/%Y")
            row[9].value = "new"
            break

    wb.save(EXCEL_FILE)
    wb.close()


def format_date_dmy(yyyymmdd: str) -> str:
    """Convert YYYYMMDD to DD/MM/YYYY for Excel consistency."""
    if re.fullmatch(r"\d{8}", yyyymmdd):
        return f"{yyyymmdd[6:8]}/{yyyymmdd[4:6]}/{yyyymmdd[:4]}"
    return yyyymmdd


# ── Ollama ──────────────────────────────────────────────────────────────────────

def call_ollama(md_content: str, criteria: str) -> dict:
    prompt = (
        f"You are a job listing scorer. Evaluate the listing below against the scoring criteria.\n"
        f"Return ONLY a valid JSON object — no explanation, no markdown.\n\n"
        f"IMPORTANT: Do NOT calculate or include a total score. "
        f"Python calculates the final score from your additions and deductions. "
        f"Your only job is to identify which criteria apply and list each one accurately "
        f"with its points value. Do not adjust points to reach a target total.\n\n"
        f"Today's date is {TODAY.strftime('%Y-%m-%d')}.\n\n"
        f"=== SCORING CRITERIA ===\n{criteria}\n========================\n\n"
        f"=== JOB LISTING ===\n{md_content}\n==================="
    )

    print(f"  Prompt size : ~{len(prompt) // 4} tokens  (window: {NUM_CTX})", flush=True)
    print(f"  Streaming response from {MODEL}:")
    print(f"  {'─'*50}")

    payload = {
        "model": MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "stream": True,
        "think": False,
        "options": {"num_ctx": NUM_CTX, "num_predict": MAX_TOKENS},
    }

    for attempt in range(1, 4):
        print(f"  ", end="", flush=True)
        full_content = ""

        resp = requests.post(OLLAMA_URL, json=payload, stream=True, timeout=300)
        resp.raise_for_status()

        for line in resp.iter_lines():
            if not line:
                continue
            try:
                chunk = json.loads(line)
            except json.JSONDecodeError:
                continue
            delta = chunk.get("message", {}).get("content", "")
            if delta:
                print(delta, end="", flush=True)
                full_content += delta
            if chunk.get("done"):
                break

        print(f"\n  {'─'*50}")

        if not full_content:
            print(f"  Empty response (attempt {attempt}) — retrying...")
            continue

        # Extract JSON from response — model may wrap it in markdown fences
        m = re.search(r'\{.*\}', full_content, re.DOTALL)
        if not m:
            print(f"  No JSON object found (attempt {attempt}): {full_content[:200]!r}")
            if attempt < 3:
                continue
            raise ValueError("No JSON found in response")

        raw_json = m.group(0)
        try:
            return json.loads(raw_json)
        except json.JSONDecodeError:
            # Fallback: escape any literal control chars inside string values
            cleaned = re.sub(r'[\x00-\x1f]', lambda c: {
                '\n': '\\n', '\t': '\\t', '\r': '\\r'
            }.get(c.group(), ''), raw_json)
            try:
                return json.loads(cleaned)
            except json.JSONDecodeError as e:
                print(f"  JSON parse error (attempt {attempt}): {e}")
                if attempt < 3:
                    print(f"  Retrying...")
                    continue
                print(f"  Raw: {full_content[:300]!r}")
                raise

    raise ValueError("Ollama returned unparseable content after 3 attempts")


# ── Console display ─────────────────────────────────────────────────────────────

def display_result(result: dict, ref_nr: int, new_stem: str) -> None:
    score = result.get("fit_score", "?")

    print(f"\n{'─'*60}")
    print(f"  Job Title  : {result.get('job_title', '?')}")
    print(f"  Company    : {result.get('company_name', '?')}")
    print(f"  Location   : {result.get('city_code', '?')}")
    print(f"  Published  : {result.get('date_published', '?')}")
    print(f"  Source     : {result.get('source', '?')}")
    print(f"\n  FIT SCORE  : {score}")

    additions = result.get("additions") or []
    if additions:
        print("\n  Additions:")
        for a in additions:
            pts = a.get("points", "")
            print(f"    + [{pts:>4}]  {a.get('criterion','')}: {a.get('detail','')}")

    deductions = result.get("deductions") or []
    if deductions:
        print("\n  Deductions:")
        for d in deductions:
            pts = d.get("points", "")
            print(f"    - [{pts:>4}]  {d.get('criterion','')}: {d.get('detail','')}")

    disqs = result.get("disqualifiers") or []
    if disqs:
        print("\n  Disqualifiers:")
        for dq in disqs:
            print(f"    ! {dq}")

    notes = (result.get("notes") or "").strip()
    if notes:
        print(f"\n  Notes      : {notes}")

    print(f"\n  ref_nr     : {ref_nr:04d}")
    print(f"  New name   : {new_stem}")
    print(f"{'─'*60}")


# ── Scoring file ────────────────────────────────────────────────────────────────

def write_scoring_file(path: Path, result: dict, ref_nr: int) -> None:
    score      = result.get("fit_score", "?")
    additions  = result.get("additions")  or []
    deductions = result.get("deductions") or []
    disqs      = result.get("disqualifiers") or []
    notes      = (result.get("notes") or "").strip()

    lines = [
        f"# Scoring Report — ref_nr {ref_nr:04d}",
        f"",
        f"| Field | Value |",
        f"|---|---|",
        f"| Job Title | {result.get('job_title', '')} |",
        f"| Company | {result.get('company_name', '')} |",
        f"| Location | {result.get('city_code', '')} |",
        f"| Published | {result.get('date_published', '')} |",
        f"| Source | {result.get('source', '')} |",
        f"| **Fit Score** | **{score}** |",
        f"",
    ]

    if additions:
        lines.append("## Additions")
        for a in additions:
            lines.append(f"- **+{a.get('points','')}** — {a.get('criterion','')}: {a.get('detail','')}")
        lines.append("")

    if deductions:
        lines.append("## Deductions")
        for d in deductions:
            lines.append(f"- **{d.get('points','')}** — {d.get('criterion','')}: {d.get('detail','')}")
        lines.append("")

    if disqs:
        lines.append("## Disqualifier Flags")
        for dq in disqs:
            lines.append(f"- {dq}")
        lines.append("")

    if notes:
        lines.append("## Notes")
        lines.append(notes)
        lines.append("")

    lines.append("## Raw JSON")
    lines.append("```json")
    lines.append(json.dumps(result, indent=2, ensure_ascii=False))
    lines.append("```")

    path.write_text("\n".join(lines), encoding="utf-8")


# ── Excel append ────────────────────────────────────────────────────────────────

def append_to_excel(result: dict, ref_nr: int, url: str) -> None:
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    date_pub = str(result.get("date_published", TODAY.strftime("%Y%m%d")))

    ref_nr_str = f"{ref_nr:04d}"
    score_str  = f"{result.get('fit_score', 0):03d}" if result.get('fit_score') is not None else None

    new_row_idx = ws.max_row + 1
    ws.append([
        ref_nr_str,                                     # A: ref_nr  (zero-padded text)
        score_str,                                      # B: fit_score (zero-padded text)
        format_date_dmy(date_pub),                      # C: date_published
        result.get("city_code", ""),                    # D: city
        result.get("client_company") or None,           # E: tag
        result.get("job_title", ""),                    # F: job_title
        result.get("company_name", ""),                 # G: company_name
        result.get("source", "LinkedIn"),               # H: source
        TODAY.strftime("%d/%m/%Y"),                     # I: date_found
        "new",                                          # J: status
        "No",                                           # K: report_produced
        None,                                           # L: processed
        url,                                            # M: url
    ])

    ws.cell(new_row_idx, 1).number_format = '@'
    ws.cell(new_row_idx, 2).number_format = '@'

    wb.save(EXCEL_FILE)
    wb.close()


# ── Per-file processing ─────────────────────────────────────────────────────────

def process_file(md_file: Path, criteria: str, idx: int, total: int, force: bool = False) -> None:
    """Score one listing, write output triplet, append Excel row. Unattended."""
    url_file = md_file.with_suffix(".url")

    print(f"\n{'='*60}")
    print(f"[{idx}/{total}]  {md_file.name}")
    print("  Calling Ollama...", flush=True)

    content = md_file.read_text(encoding="utf-8")
    url = extract_url(content)

    if "no longer accepting applications" in content.lower() and not force:
        print(f"  → CLOSED (no longer accepting applications) — skipping.")
        closed_dir = INPUT_DIR / "closed"
        closed_dir.mkdir(exist_ok=True)
        shutil.move(str(md_file), closed_dir / md_file.name)
        if url_file.exists():
            shutil.move(str(url_file), closed_dir / url_file.name)
        return

    try:
        result = call_ollama(content, criteria)
    except requests.exceptions.ConnectionError:
        print(f"  ERROR: cannot reach Ollama at {OLLAMA_URL}")
        print("  Skipping this file.")
        return
    except Exception as e:
        print(f"  ERROR: {e}")
        print("  Skipping this file.")
        return

    # ── Normalise result fields ─────────────────────────────────────────────────

    # Recalculate fit_score from the breakdown — LLMs are unreliable at arithmetic.
    # Additions: industry bonus applies only once regardless of how many matched.
    additions  = result.get("additions")  or []
    deductions = result.get("deductions") or []

    industry_bonus_applied = False
    additions_total = 0
    for a in additions:
        try:
            pts = abs(int(a.get("points", 0)))  # always positive regardless of LLM sign
        except (ValueError, TypeError):
            pts = 0
        crit = str(a.get("criterion", "")).lower()
        if "industr" in crit:
            if industry_bonus_applied:
                a["points"] = 0
                a["detail"] = a.get("detail", "") + " [duplicate — industry bonus applied once only]"
                continue
            industry_bonus_applied = True
        a["points"] = pts
        additions_total += pts

    deductions_total = 0
    for d in deductions:
        try:
            pts = -abs(int(d.get("points", 0)))  # always negative regardless of LLM sign
        except (ValueError, TypeError):
            pts = 0
        d["points"] = pts
        deductions_total += pts

    score = max(0, 100 + additions_total + deductions_total)
    result["fit_score"] = score
    result["additions"]  = additions
    result["deductions"] = deductions

    date_pub = str(result.get("date_published", "")).strip()
    if not re.fullmatch(r"\d{8}", date_pub):
        date_pub = TODAY.strftime("%Y%m%d")
    result["date_published"] = date_pub

    # Use LLM city_code if valid, otherwise derive it
    raw_city = str(result.get("city_code", "")).strip()
    if not re.fullmatch(r"[A-Za-z]{1,7}", raw_city):
        raw_city = city_code(raw_city)
    result["city_code"] = raw_city.upper()

    result.setdefault("source", "LinkedIn")

    # ── Build new filename stem ─────────────────────────────────────────────────

    if force:
        existing_ref = lookup_ref_nr_by_url(url)
    else:
        existing_ref = None

    if existing_ref is not None:
        ref_nr = existing_ref
        print(f"  Force-rescore: reusing ref_nr {ref_nr:04d}")
    else:
        ref_nr = next_ref_nr()

    title_s        = sanitize(result.get("job_title", ""))
    company_s      = sanitize(result.get("company_name", ""))
    src_s          = sanitize(result.get("source", "LinkedIn"))
    client_company = sanitize(result.get("client_company", ""))

    new_stem = f"{ref_nr:04d}_{score:03d}_{date_pub}_{result['city_code']}_{title_s}_{company_s}_{src_s}"
    if client_company:
        new_stem += f"__{client_company}"

    display_result(result, ref_nr, new_stem)

    # ── Commit ──────────────────────────────────────────────────────────────────

    if existing_ref is not None:
        update_excel_row_for_rescore(ref_nr, result)
    else:
        append_to_excel(result, ref_nr, url)

    OUTPUT_DIR.mkdir(exist_ok=True)
    new_md  = OUTPUT_DIR / f"{new_stem}.md"
    new_url = OUTPUT_DIR / f"{new_stem}.url"

    # Resolve collisions
    counter = 1
    while new_md.exists() or new_url.exists():
        new_md  = OUTPUT_DIR / f"{new_stem}_{counter:02d}.md"
        new_url = OUTPUT_DIR / f"{new_stem}_{counter:02d}.url"
        counter += 1

    shutil.move(str(md_file), new_md)
    if url_file.exists():
        shutil.move(str(url_file), new_url)

    scoring_file = OUTPUT_DIR / f"{new_stem}_SCORING.md"
    write_scoring_file(scoring_file, result, ref_nr)

    print(f"  Committed  : ref_nr {ref_nr:04d}  score {score:03d}  →  {new_md.name}")


# ── Main ────────────────────────────────────────────────────────────────────────

def main():
    force = "--force" in sys.argv

    if _load_model(5) != MODEL:
        _unload_all_models()

    md_files = sorted(f for f in INPUT_DIR.glob("*.md")
                      if not f.stem.endswith("_REPORT")
                      and not f.stem.endswith("_SCORING"))
    if not md_files:
        print(f"{INPUT_DIR.name}/ has no .md files — nothing to do.")
        return

    if not CRITERIA.exists():
        print(f"ERROR: {CRITERIA} not found.")
        return

    criteria = CRITERIA.read_text(encoding="utf-8")

    print(f"Stage 06 — LLM Scoring")
    print(f"  Model      : {MODEL}")
    print(f"  Input      : {INPUT_DIR.name}/  ({len(md_files)} listings)")
    print(f"  Output     : {OUTPUT_DIR.name}/")
    print(f"  Excel      : {EXCEL_FILE.name}")
    if force:
        print(f"  Force mode : closed-listing check bypassed; ref_nr reused if URL known")
    print(f"  Next ref_nr: {next_ref_nr():04d}")

    for i, md_file in enumerate(md_files, 1):
        process_file(md_file, criteria, i, len(md_files), force=force)

    print("\nDone.")


if __name__ == "__main__":
    main()
