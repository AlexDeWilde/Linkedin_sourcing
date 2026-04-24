import sys, shutil, re, time, threading, webbrowser, json
from datetime import date
from pathlib import Path
from flask import Flask, jsonify, request
from openpyxl import load_workbook

sys.stdout.reconfigure(write_through=True)

BASE_DIR   = Path(__file__).parent
EXCEL_FILE = BASE_DIR / '06-listings_db.xlsx'

FOLDERS = {
    'new':        '06-LLM_scored',
    'rejects':    '07-rejects',
    'consider':   '08-consider',
    'priorities': '09-priorities',
    'in_process': '10-in_process',
    'applied':    '11-applied',
}
COLUMN_ORDER = ['new', 'consider', 'priorities', 'in_process']

# Excel column headers — first 13 match the existing schema, 14-20 are stage 07 additions
EXCEL_HEADERS = [
    'ref_nr', 'fit_score', 'date_published', 'city', 'tag',
    'job_title', 'company_name', 'source', 'date_found', 'status',
    'report_produced', 'processed', 'url',
    'date_considered', 'date_priorities', 'date_in_process',
    'date_applied', 'date_rejected', 'comment',
    'reject_reason',
]

# Status value written to Excel per folder key
STATUS_FOR_COL = {
    'new':        'new',
    'rejects':    'rejected',
    'consider':   'in_consideration',
    'priorities': 'priorities',
    'in_process': 'in_process',
    'applied':    'applied',
}

# Which date column to stamp when a card moves to each folder
DATE_COL_FOR = {
    'consider':   'date_considered',
    'priorities': 'date_priorities',
    'in_process': 'date_in_process',
    'applied':    'date_applied',
    'rejects':    'date_rejected',
}

app = Flask(__name__)


# ── Excel helpers ────────────────────────────────────────────────────────────

def setup_excel_columns():
    """Ensure the Excel file has all required column headers in row 1."""
    if not EXCEL_FILE.exists():
        return
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    if ws.max_row == 0:
        wb.close()
        return

    first_val = str(ws.cell(1, 1).value or '').strip().lower()
    if first_val != 'ref_nr':
        # No header row — insert one at the top
        ws.insert_rows(1)
        for i, name in enumerate(EXCEL_HEADERS[:13], 1):
            ws.cell(1, i).value = name

    # Add any missing new columns to the header row
    existing = {str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1)}
    col_idx = ws.max_column + 1
    for name in EXCEL_HEADERS:
        if name not in existing:
            ws.cell(1, col_idx).value = name
            col_idx += 1

    wb.save(EXCEL_FILE)
    wb.close()


def get_col_map(ws) -> dict:
    """Return {header_name: column_number (1-indexed)} from the header row."""
    return {str(cell.value).strip(): cell.column
            for cell in ws[1] if cell.value}


def excel_update_row(ref_nr_str: str, updates: dict) -> bool:
    """Update named columns for the row matching ref_nr. Returns True if found."""
    if not EXCEL_FILE.exists():
        return False
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        col_map = get_col_map(ws)
        ref = ref_nr_str.zfill(4)

        for row in ws.iter_rows(min_row=2):
            if str(row[0].value or '').strip().zfill(4) == ref:
                for col_name, value in updates.items():
                    if col_name in col_map:
                        ws.cell(row[0].row, col_map[col_name]).value = value
                wb.save(EXCEL_FILE)
                wb.close()
                return True

        wb.close()
        return False
    except Exception as e:
        print(f'  Excel update warning: {e}')
        return False


# ── File / listing helpers ───────────────────────────────────────────────────

def col_path(col: str) -> Path:
    return BASE_DIR / FOLDERS[col]


def parse_meta(scoring_path: Path) -> dict:
    text = scoring_path.read_text(encoding='utf-8', errors='replace')

    def grab(label):
        pattern = r'\|\s*\*{0,2}' + re.escape(label) + r'\*{0,2}\s*\|\s*\*{0,2}([^|\n]+?)\*{0,2}\s*\|'
        m = re.search(pattern, text, re.IGNORECASE)
        return m.group(1).strip() if m else ''

    return {
        'job_title': grab('Job Title'),
        'company':   grab('Company'),
        'location':  grab('Location'),
        'published': grab('Published'),
        'fit_score': grab('Fit Score'),
    }


def stem_of(scoring_path: Path) -> str:
    return scoring_path.name[:-len('_SCORING.md')]


def listings_in(col: str) -> list:
    p = col_path(col)
    if not p.exists():
        return []
    results = []
    for sf in sorted(p.glob('*_SCORING.md')):
        stem = stem_of(sf)
        meta = parse_meta(sf)
        results.append({
            'stem':        stem,
            'col':         col,
            'job_title':   meta['job_title'],
            'company':     meta['company'],
            'location':    meta['location'],
            'published':   meta['published'],
            'fit_score':   meta['fit_score'],
            'has_comment': (p / f'{stem}_comment.md').exists(),
        })
    return results


def read_url(path: Path) -> str:
    if not path.exists():
        return ''
    text = path.read_text(encoding='utf-8', errors='replace')
    m = re.search(r'URL=(.+)', text)
    return m.group(1).strip() if m else ''


# ── Routes ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return (BASE_DIR / '07-review.html').read_text(encoding='utf-8')


@app.route('/api/listings')
def api_listings():
    return jsonify({col: listings_in(col) for col in COLUMN_ORDER})


@app.route('/api/content', methods=['POST'])
def api_content():
    data = request.json
    stem, col = data['stem'], data['col']
    p = col_path(col)

    def read(path):
        return path.read_text(encoding='utf-8', errors='replace') if path.exists() else ''

    return jsonify({
        'scoring': read(p / f'{stem}_SCORING.md'),
        'jd':      read(p / f'{stem}.md'),
        'comment': read(p / f'{stem}_comment.md'),
        'url':     read_url(p / f'{stem}.url'),
    })


@app.route('/api/save_comment', methods=['POST'])
def api_save_comment():
    data = request.json
    stem, col, text = data['stem'], data['col'], data['text']
    p = col_path(col)
    (p / f'{stem}_comment.md').write_text(text, encoding='utf-8')
    excel_update_row(stem[:4], {'comment': text.strip()})
    return jsonify({'ok': True})


@app.route('/api/move', methods=['POST'])
def api_move():
    data = request.json
    stem, from_col, to_col = data['stem'], data['from'], data['to']
    reason = (data.get('reason') or '').strip()

    if from_col not in FOLDERS or to_col not in FOLDERS:
        return jsonify({'ok': False, 'error': 'Unknown column'}), 400

    src = col_path(from_col)
    dst = col_path(to_col)
    dst.mkdir(exist_ok=True)

    # Move every file sharing the stem (catches sidecars like _REPORT.md, _comment.md, etc.)
    # The "stem." + "stem_" split avoids matching stems that are prefixes of others.
    for f in list(src.glob(f'{stem}.*')) + list(src.glob(f'{stem}_*')):
        shutil.move(str(f), str(dst / f.name))

    # Update Excel: status + stage date (+ reject_reason when rejecting)
    today = date.today().strftime('%d/%m/%Y')
    updates = {'status': STATUS_FOR_COL.get(to_col, to_col)}
    if to_col in DATE_COL_FOR:
        updates[DATE_COL_FOR[to_col]] = today
    if to_col == 'rejects' and reason:
        updates['reject_reason'] = reason
    excel_update_row(stem[:4], updates)

    return jsonify({'ok': True})


def sync_all_to_excel() -> dict:
    """Scan all folders and update Excel status, stage dates, and comments.
    Returns {'ok': bool, 'updated': int, 'not_found': int, 'error'?: str}."""
    today = date.today().strftime('%d/%m/%Y')

    all_updates: dict[str, dict] = {}
    for col, folder_name in FOLDERS.items():
        folder = BASE_DIR / folder_name
        if not folder.exists():
            continue
        for sf in sorted(folder.glob('*_SCORING.md')):
            stem   = stem_of(sf)
            ref_nr = stem[:4]
            upd    = {'status': STATUS_FOR_COL.get(col, col)}
            if col in DATE_COL_FOR:
                upd[DATE_COL_FOR[col]] = today
            comment_f = folder / f'{stem}_comment.md'
            if comment_f.exists():
                upd['comment'] = comment_f.read_text(encoding='utf-8', errors='replace').strip()
            all_updates[ref_nr] = upd

    if not all_updates or not EXCEL_FILE.exists():
        return {'ok': True, 'updated': 0, 'not_found': 0}

    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        col_map = get_col_map(ws)
        updated, remaining = 0, dict(all_updates)

        for row in ws.iter_rows(min_row=2):
            ref = str(row[0].value or '').strip().zfill(4)
            if ref in remaining:
                for col_name, value in remaining[ref].items():
                    if col_name in col_map:
                        ws.cell(row[0].row, col_map[col_name]).value = value
                updated += 1
                del remaining[ref]

        wb.save(EXCEL_FILE)
        wb.close()
        return {'ok': True, 'updated': updated, 'not_found': len(remaining)}
    except Exception as e:
        return {'ok': False, 'error': str(e), 'updated': 0, 'not_found': 0}


@app.route('/api/sync_all', methods=['POST'])
def api_sync_all():
    result = sync_all_to_excel()
    if not result.get('ok'):
        return jsonify(result), 500
    return jsonify(result)


# ── Score editing ─────────────────────────────────────────────────────────────

def recalculate_score(result: dict) -> int:
    """Normalise points (additions +, deductions -) in-place and return final score."""
    additions  = result.get('additions')  or []
    deductions = result.get('deductions') or []

    industry_bonus_applied = False
    additions_total = 0
    for a in additions:
        try:
            pts = abs(int(a.get('points', 0)))
        except (ValueError, TypeError):
            pts = 0
        crit = str(a.get('criterion', '')).lower()
        if 'industr' in crit:
            if industry_bonus_applied:
                a['points'] = 0
                continue
            industry_bonus_applied = True
        a['points'] = pts
        additions_total += pts

    deductions_total = 0
    for d in deductions:
        try:
            pts = -abs(int(d.get('points', 0)))
        except (ValueError, TypeError):
            pts = 0
        d['points'] = pts
        deductions_total += pts

    return max(0, 100 + additions_total + deductions_total)


def render_scoring_md(result: dict, ref_nr: int) -> str:
    score      = result.get('fit_score', 0)
    additions  = result.get('additions')  or []
    deductions = result.get('deductions') or []
    disqs      = result.get('disqualifiers') or []
    notes      = (result.get('notes') or '').strip()

    lines = [
        f'# Scoring Report — ref_nr {ref_nr:04d}', '',
        '| Field | Value |', '|---|---|',
        f"| Job Title | {result.get('job_title', '')} |",
        f"| Company | {result.get('company_name', '')} |",
        f"| Location | {result.get('city_code', '')} |",
        f"| Published | {result.get('date_published', '')} |",
        f"| Source | {result.get('source', '')} |",
        f'| **Fit Score** | **{score}** |', '',
    ]
    if additions:
        lines.append('## Additions')
        for a in additions:
            lines.append(f"- **+{a.get('points','')}** — {a.get('criterion','')}: {a.get('detail','')}")
        lines.append('')
    if deductions:
        lines.append('## Deductions')
        for d in deductions:
            lines.append(f"- **{d.get('points','')}** — {d.get('criterion','')}: {d.get('detail','')}")
        lines.append('')
    if disqs:
        lines.append('## Disqualifier Flags')
        for dq in disqs:
            lines.append(f'- {dq}')
        lines.append('')
    if notes:
        lines.append('## Notes')
        lines.append(notes)
        lines.append('')
    lines += ['## Raw JSON', '```json',
              json.dumps(result, indent=2, ensure_ascii=False), '```']
    return '\n'.join(lines)


@app.route('/api/save_scoring', methods=['POST'])
def api_save_scoring():
    data    = request.json
    stem    = data['stem']
    col     = data['col']
    content = data['content']
    p       = col_path(col)

    try:
        result = json.loads(content)
    except json.JSONDecodeError as e:
        return jsonify({'ok': False, 'error': f'Invalid JSON: {e}'}), 400

    new_score           = recalculate_score(result)
    result['fit_score'] = new_score
    ref_nr_str          = stem[:4]
    new_stem            = f'{ref_nr_str}_{new_score:03d}_{stem[9:]}'

    if new_stem != stem:
        # Rename every file sharing the stem (catches sidecars like _REPORT.md).
        for old_f in list(p.glob(f'{stem}.*')) + list(p.glob(f'{stem}_*')):
            new_f = p / (new_stem + old_f.name[len(stem):])
            old_f.rename(new_f)

    (p / f'{new_stem}_SCORING.md').write_text(
        render_scoring_md(result, int(ref_nr_str)), encoding='utf-8')

    try:
        excel_update_row(ref_nr_str, {'fit_score': f'{new_score:03d}'})
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Excel update failed: {e}'}), 500

    return jsonify({'ok': True, 'new_stem': new_stem, 'new_score': new_score})


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == '__main__':
    setup_excel_columns()

    sync_result = sync_all_to_excel()
    if sync_result.get('ok'):
        print(f"Startup Sync DB: {sync_result['updated']} row(s) updated, "
              f"{sync_result['not_found']} not found in Excel.")
    else:
        print(f"Startup Sync DB warning: {sync_result.get('error')}")

    def open_browser():
        time.sleep(1.2)
        webbrowser.open('http://localhost:5000')

    threading.Thread(target=open_browser, daemon=True).start()
    print('Job Review Console -> http://localhost:5000')
    print('Press Ctrl+C to stop.')
    app.run(host='127.0.0.1', port=5000, debug=False)
